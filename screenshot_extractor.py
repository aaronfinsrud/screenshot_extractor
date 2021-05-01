import sys
import zipfile
import shutil
import os
import tempfile
import re
import openpyxl
from tkinter import *
from tkinter.filedialog import askopenfilename

#REFERENCE: https://stackoverflow.com/questions/57622795/extract-image-from-excel-in-nodejs

##TODO:


##MAIN##
#helper function
def quicksort(array):
  if len(array) < 2:
    # base case, arrays with 0 or 1 element are already "sorted"
    return array
  else:
    # recursive case
    pivot = array[0]
    # sub-array of all the elements less than the pivot
    less = [i for i in array[1:] if i <= pivot]
    # sub-array of all the elements greater than the pivot
    greater = [i for i in array[1:] if i > pivot]
    return quicksort(less) + [pivot] + quicksort(greater)

#prevents command line from exiting upon error
def show_exception_and_exit(exc_type, exc_value, tb):
    import traceback
    traceback.print_exception(exc_type, exc_value, tb)
    input("Press <return> to exit...")
    sys.exit(-1)
sys.excepthook = show_exception_and_exit
##MAC
root = Tk()
root.update()
root.withdraw()
wb_path = askopenfilename()
##

if len(sys.argv) > 0:
    wb_path = wb_path
    print(wb_path)
    print()
else:
    raise Exception('No file selected. Please close this window and run the script again.')

wb_folder =  os.path.dirname(wb_path)
# creates a temporary directory #
temp_directory = tempfile.mkdtemp()
shutil.copyfile(wb_path, os.path.join(temp_directory,"copy.zip"))
temp_screenshot_directory = os.path.join(wb_folder,"screenshots")
# creates folder to hold screenshots, if non-existent
if not os.path.exists(temp_screenshot_directory):
    os.mkdir(os.path.join(wb_folder,"screenshots"))
with zipfile.ZipFile(os.path.join(temp_directory,"copy.zip"), 'r') as zip_ref:
    zip_ref.extractall(temp_screenshot_directory)

image_IDs = {}
#STEP 1:
xml_rels_dir = os.path.join(temp_screenshot_directory,"xl","drawings","_rels")

# sorts the files alphabetically
file_list = []
for rels_filename in os.listdir(xml_rels_dir):
    file_list.append(rels_filename)
file_list = sorted(file_list)   

# each rels_filename corresponds to a sheet
for k, rels_filename in enumerate(file_list):
    file = open(os.path.join(xml_rels_dir,rels_filename), "r")
    rels_contents = file.read()
    file.close()
    filename = re.findall(r'[T][a][r][g][e][t][=]\"(.+?)\"',rels_contents)
    file_ID = re.findall(r'[I][d][=]\"(.+?)\"',rels_contents)
    filename = [i.split(r"/")[-1] for i in filename]
    # sheet#:file_ID:filename
    image_IDs[k] = dict(zip(file_ID, filename))


sheets = []

#STEP 2:

xml_files_dir = os.path.join(temp_screenshot_directory,"xl","drawings")
# sorts the files alphabetically
file_list = []
for f in os.listdir(xml_files_dir):
    file_list.append(f)
file_list = sorted(file_list)  
xml_files = []
# iterates through each xml file in ".zip\xl\drawings"
for f in file_list:
    name, ext = os.path.splitext(f)
    if ext == '.xml':
        xml_files.append(f)
# reads and splits contents of each xml file
for j, xml_filename in enumerate(xml_files):
    print(xml_filename)
    file = open(os.path.join(xml_files_dir,xml_filename), "r")
    xml_contents = file.read()
    file.close()
    d = r'([r][:][e][m][b][e][d][=]\".+?\")'
    xml_contents_array = re.split(d,xml_contents)
    xml_contents_array[0:-2]
    xml_contents_array = [ x+y for x,y in zip(xml_contents_array[0::2], xml_contents_array[1::2]) ]
    image_filenames = {}
    # iterates over content blocks (from xml) correspdonding to an image linkage
    for m, content_block in enumerate(xml_contents_array):
        d = r'([r][:][e][m][b][e][d][=]\"(.+?)\")'
        fileID = re.search(d,content_block).group(2)
        # gets filename of image correspdonding to content_block
        filename = image_IDs[j].get(fileID)
        print(filename)
        # gets all rows with images
        row_range = re.findall(r'<xdr:row>(.+?)\</xdr:row>',content_block)
        #converts row range data to integers
        row_range = [int(i) for i in row_range]
        print("row range")
        print(row_range)      
        # image_filenames dict with key as filename and value as row range     
        if filename in image_filenames:
            image_filenames[filename].append(row_range)
        else:
            image_filenames[filename] = []
            image_filenames[filename].append(row_range)

    # appends image_filenames dict to sheets list
    sheets.append(image_filenames) 
print("sheets")
print(sheets)

## STEP 2.5:
# iterates through .rels files in zip\xl\worksheets\_rels
xml_rels_files_dir = os.path.join(temp_screenshot_directory,"xl","worksheets","_rels")
sheet_indexes = []

# sorts the files alphabetically
file_list = []
for f in os.listdir(xml_rels_files_dir):
    file_list.append(f)
file_list = sorted(file_list)  

for f in file_list:
    print(f)
    f = str(f)
    value = re.search(r'\d+',f).group(0)
    value = int(value)
    sheet_indexes.append(value)
sheet_indexes = quicksort(sheet_indexes)
print(sheet_indexes)


## STEP 3:
image_fn = {}
keys_list = []
# helper function to extract values from excel spreadsheet
def get_cell_range(start_col, start_row, end_col, end_row):
    key_list = []
    ##column = worksheet.col(end_col)##
    if start_row == end_row:
      end_row = end_row+1
    if start_row == 0:
      start_row = start_row +1
      end_row = end_row +1
    if end_row-start_row >= 2:
      end_row = end_row+1
    for n in range(start_row, end_row):    
        key = worksheet.cell(row=n, column=end_col).value####(column[n]).value
        key_list.append(key)
    return key_list
## update
## workbook = xlrd.open_workbook(wb_path)
## update
workbook = openpyxl.load_workbook(wb_path)
##

# iterates over sheets list
for t, sheet in enumerate (sheets):
    image_filenames = sheet
    # iterates over image_filenames dict (fn:row ranges)
    for key, value in image_filenames.items():
        sheet_num = sheet_indexes[t]
        print("sheet number")
        print(sheet_num)
        ## update
        ## worksheet = workbook.sheet_by_index(sheet_num-1)
        ##update
        worksheet = workbook.worksheets[sheet_num-1] ## double check 
        ##
        row_range = value
        key_list = []
        # identifies column with key header
        column_headers = []
        key_column_header = ["key","keys","Keys","Key"]
        for r in [1,1]:
            for c in range(worksheet.max_column):##
                c = c+1
                cell = worksheet.cell(row=r, column=c)#
                
                # checks for key column
                for val in key_column_header:
                    if cell.value and cell.value.strip() == val:
                        
                        start_col = c
                        end_col = c  
        print("column")
        print(end_col)
        print("rows")
        print(row_range)
        key_list1 = []
        for rrange in row_range:
            key_list = get_cell_range(start_col, min(rrange)+1, end_col, max(rrange)+1)
            key_list1.extend(key_list)
        print("kl: ")
        print(key_list1)
        print()
        # makes keys list callable by filename
        # if key exists,append to existing
        if key in image_fn:
            image_fn[key].extend(key_list1) 
        else:
            image_fn[key] = key_list1
        

saved_files = []
t = 1
# STEP: 4

# sorts the files alphabetically
file_list = []
for f in os.listdir(os.path.join(temp_screenshot_directory, "xl","media")):
    file_list.append(f)
file_list = sorted(file_list) 


# iterates over each screenshot within zip
for filename in file_list:
    if filename.endswith(".jpg") or filename.endswith(".png") or filename.endswith(".jpeg"):
        # screenshot_file is the file to be copied
        screenshot_file = os.path.join(temp_screenshot_directory, "xl","media", filename)
        file_name, file_extension = os.path.splitext(screenshot_file)
        for key, value in image_fn.items():
            if filename == key:
                for cell_value in value:
                    # incase the cell_value has already been used 
                    # incase the cell_value is empty
                    if cell_value == "":
                        cell_value = "no_key_"+str(t)
                        t += 1
                    if cell_value in saved_files:
                        shutil.rmtree(temp_screenshot_directory)
                        ## MAC  
                        root.destroy()
                        ##
                        print()
                        print("------------------------------------")
                        print("ERROR!")
                        print("Duplicate key: ")
                        print(cell_value)
                        print()
                    
                        print("Make sure:")
                        print("- Each key is unique")
                        print("- Each key aligns with no more than 1 screenshot")
                        print("------------------------------------")
                        print()
                        raise Exception("DUPLICATE KEY")
                    if not isinstance(cell_value, str) and cell_value != int(cell_value):
                      cell_value = str(cell_value)
                    elif not isinstance(cell_value, str) and cell_value == int(cell_value):
                      cell_value = str(int(cell_value))
                    else:
                      cell_value
                    ## handling for filenames that are too long or contain illegal chars
                    invalid = '<>:"/\|?* '
                    for char in invalid:
                      if char in cell_value:
                        print()
                        print("------------------------------------")
                        print("ERROR!")
                        print("Illegal character in key: ")
                        print(cell_value)
                        print()
                        print("Make sure keys do not contain any of the following characters:")
                        print('<>:"/\|?* ')
                        print("------------------------------------")
                        print()
                        shutil.rmtree(temp_screenshot_directory)
                        ## MAC  
                        root.destroy()
                        ##
                        raise Exception("ILLEGAL CHAR IN KEY")
                        
                      elif len(cell_value+file_extension) > 260:
                        print()
                        print("------------------------------------")
                        print("ERROR!")
                        print("Key length limit exceeded: ")
                        print(cell_value)
                        print()
                        print("Make sure keys are less than 255 characters in length")
                        print("------------------------------------")
                        print()
                        shutil.rmtree(temp_screenshot_directory)
                        ## MAC  
                        root.destroy()
                        ##
                        raise Exception("KEY LENGTH LIMIT EXCEEDED")
                      
                    shutil.copyfile(screenshot_file, os.path.join(temp_screenshot_directory,"xl","media",cell_value+file_extension))
                    saved_files.append(cell_value)
    else:
        continue
    os.remove(os.path.join(temp_screenshot_directory, "xl","media",filename))
# saves screenshots to a .zip within temp screenshot directory
new_directory = os.path.join(temp_screenshot_directory, "xl","media")
file_name, file_extension = os.path.splitext(wb_path)
shutil.make_archive(file_name+"_screenshots", 'zip',new_directory)

# deletes the temp_screenshot_directory
shutil.rmtree(temp_screenshot_directory)
root.destroy()
